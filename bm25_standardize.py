import re, sys
import pandas as pd
import numpy as np
from rank_bm25 import BM25Okapi
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os

# ── CONFIG ────────────────────────────────────────────────────────────────────
SIMILARITY_THRESHOLD = 0.60    # A score >= 60% of the max possible score is a confident match. Tune this if needed.

# ── TOKENIZER ─────────────────────────────────────────────────────────────────
# Split on commas, spaces, #, /, -, (, ) so each meaningful token is separate.
# Example: "CBL,WIRE,UL1007,22#(7/30TA),BLK,OD2.0,ALPHA#3055"
#       -> ["CBL","WIRE","UL1007","22","7","30TA","BLK","OD2.0","ALPHA","3055"]
def tokenize(text):
    text = str(text).upper().strip()
    text = (text.replace("\uff0c", ",").replace("\uff08", "(")
                .replace("\uff09", ")").replace("\u3000", " "))
    tokens = re.split(r"[,\s#/()\-]+", text)
    return [t for t in tokens if t]

# POST-CORRECTION 
COLOR_TOKENS = [
    "GRN/YEL-ST","GRN/YEL-RI","WHT/GRY-ST","GN/YW-ST",
    "ORG/WHT","ORG/BLK","RED/WHT","RED/BLK","WHT/BLU","BLU/WHT",
    "WHT/BLK","WHT/RED","WHT/GRN","WHT/GRY","WHT/BRN",
    "YEL/GRN","YEL/WHT","YEL/RED","GRN/YEL","PINK/WHT",
    "LT-BLU","LT-GRN","DK-BLU","DK-GRN","D-BLU","D-GRN",
    "BEIGE","SLATE","DARK","PINK",
    "BLK","RED","WHT","BLU","GRN","YEL","ORG","BRN","GRY","PUR","PNK","VIO",
]
COLOR_PAT = "|".join(re.escape(c) for c in COLOR_TOKENS)

def fix_fullwidth(s):
    return (s.replace("\uff0c",",").replace("\uff08","(")
             .replace("\uff09",")").replace("\u3000"," ")
             .replace("，",",").replace("（","(").replace("）",")"))

def fix_comma_before_color(s):
    return re.sub(r"\)(" + COLOR_PAT + r")([,\s]|$)",
                  lambda m: ")," + m.group(1) + m.group(2), s)

def fix_comma_before_od(s):
    return re.sub(r"(?<!\))([A-Za-z])OD(\d)",
                  lambda m: m.group(1) + ",OD" + m.group(2), s)

def fix_comma_after_od(s):
    return re.sub(r"(OD[\d\.]+)([A-Za-z#])",
                  lambda m: m.group(1) + "," + m.group(2), s)

def fix_space_in_od(s):
    return re.sub(r"\bOD\s+(\d)", r"OD\1", s)

def fix_whitespace(s):
    return re.sub(r"\s+,", ",", s).strip()

def post_correct(desc):
    changes = []
    def apply(fn, label):
        nonlocal desc
        s = fn(desc)
        if s != desc:
            changes.append(label)
            desc = s
    apply(fix_fullwidth,          "full-width chars")
    apply(fix_comma_before_color, "comma before color")
    apply(fix_comma_before_od,    "comma before OD")
    apply(fix_comma_after_od,     "comma after OD")
    apply(fix_space_in_od,        "space in OD")
    apply(fix_whitespace,         "whitespace")
    return desc, "; ".join(changes) if changes else ""


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main(good_path, nogood_path, output_path):

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print("Loading data...")
    good_df = pd.read_excel(good_path,   sheet_name="Raw")
    nogood_df = pd.read_excel(nogood_path, sheet_name="Raw")
    good_df.columns = good_df.columns.str.strip()
    nogood_df.columns = nogood_df.columns.str.strip()

    good_descs = good_df["Full Item Description"].astype(str).tolist()
    nogood_descs = nogood_df["Full Item Description"].astype(str).tolist()
    good_parts = good_df["Part Number"].astype(str).tolist()
    print(f"  Good   : {len(good_descs)}")
    print(f"  No Good: {len(nogood_descs)}")

    # ── Build BM25 index ──────────────────────────────────────────────────────
    print("\nBuilding BM25 index on Good descriptions...")
    tokenized_good = [tokenize(d) for d in good_descs]
    bm25 = BM25Okapi(tokenized_good)
    print(f"  Index built with {len(tokenized_good)} documents")

    # ── Score & match ─────────────────────────────────────────────────────────
    print("\nMatching...")
    results = []
    BATCH = 500

    for start in range(0, len(nogood_descs), BATCH):
        end = min(start + BATCH, len(nogood_descs))
        batch = nogood_descs[start:end]

        for i, original in enumerate(batch):
            part_no = nogood_df["Part Number"].iloc[start + i]
            query_tokens = tokenize(original)

            scores = bm25.get_scores(query_tokens)
            best_idx = int(np.argmax(scores))
            best_raw = float(scores[best_idx])

            # Normalize score: express as % of the self-score of the best match
            # so the scale becomes roughly [0, 1] for easy comparison with TF-IDF
            self_scores = bm25.get_scores(tokenized_good[best_idx])
            self_max = float(self_scores[best_idx])
            norm_score = min((best_raw / self_max) if self_max > 0 else 0.0, 1.0)

            ml_match        = good_descs[best_idx]
            matched_good_pn = good_parts[best_idx]

            if norm_score >= SIMILARITY_THRESHOLD:
                final, corrections = post_correct(ml_match)
                status = "AUTO"
            else:
                final       = ""
                corrections = ""
                status      = "MANUAL REVIEW NEEDED"

            results.append({
                "No_Good Part#"           : part_no,
                "Original Description"    : original,
                "Standardized Description": final,
                "Status"                  : status,
                "BM25 Score (normalized)" : round(norm_score, 4),
                "Matched Good Part#"      : matched_good_pn,
                "BM25 Match (reference)"  : ml_match,
                "Post-corrections"        : corrections if corrections else "none",
            })

        print(f"  {end}/{len(nogood_descs)}...", end="\r")

    print("\nDone matching.")

    # ── Write Excel ───────────────────────────────────────────────────────────
    print("Writing Excel...")
    out_df = pd.DataFrame(results)
    out_df.to_excel(output_path, index=False, sheet_name="Results")

    wb = load_workbook(output_path)
    ws = wb["Results"]

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    green = PatternFill("solid", fgColor="E2EFDA")
    red   = PatternFill("solid", fgColor="F4CCCC")
    for row in ws.iter_rows(min_row=2):
        fill = green if row[3].value == "AUTO" else red
        for cell in row:
            cell.fill = fill

    for idx, w in enumerate([15, 65, 65, 22, 18, 18, 65, 35], 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    wb.save(output_path)

    #SUMMARY
    auto   = (out_df["Status"] == "AUTO").sum()
    manual = (out_df["Status"] == "MANUAL REVIEW NEEDED").sum()
    total  = len(out_df)
    print(f"\n{'='*40}")
    print(f"Total processed: {total}")
    print(f"Auto standardized: {auto}  ({auto/total*100:.1f}%)")
    print(f"Manual review: {manual}  ({manual/total*100:.1f}%)")
    print(f"Threshold used: {SIMILARITY_THRESHOLD}")
    print(f"Output: {output_path}")

if __name__ == "__main__":
    good_path   = sys.argv[1] if len(sys.argv) > 1 else "data/1.Good.xlsx"
    nogood_path = sys.argv[2] if len(sys.argv) > 2 else "data/2.No_Good.xlsx"
    output_path = sys.argv[3] if len(sys.argv) > 3 else "outputs/BM25_Standardized_Output.xlsx"
    main(good_path, nogood_path, output_path)