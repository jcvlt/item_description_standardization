import re, sys
import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SIMILARITY_THRESHOLD = 0.95

# Color tokens
COLOR_TOKENS = [
    "GRN/YEL-ST","GRN/YEL-RI","WHT/GRY-ST","GN/YW-ST",
    "ORG/WHT","ORG/BLK","RED/WHT","RED/BLK","WHT/BLU","BLU/WHT",
    "WHT/BLK","WHT/RED","WHT/GRN","WHT/GRY","WHT/BRN",
    "YEL/GRN","YEL/WHT","YEL/RED","GRN/YEL","PINK/WHT",
    "LT-BLU","LT-GRN","DK-BLU","DK-GRN","D-BLU","D-GRN",
    "BEIGE","SLATE","DARK","PINK",
    "BLK","RED","WHT","BLU","GRN","YEL","ORG","BRN","GRY","PUR","PNK","VIO",
]
COLOR_PAT = "|".join(re.escape(c) for c in COLOR_TOKENS)

# Post-correction rules
def fix_fullwidth(s):
    return (s.replace("\uff0c",",").replace("\uff08","(")
             .replace("\uff09",")").replace("\u3000"," ")
             .replace("，",",").replace("（","(").replace("）",")"))

def fix_comma_before_color(s):
    return re.sub(r'\)(' + COLOR_PAT + r')([,\s]|$)',
                  lambda m: ")," + m.group(1) + m.group(2), s)

def fix_comma_before_od(s):
    return re.sub(r'(?<!\))([A-Za-z])OD(\d)',
                  lambda m: m.group(1) + ",OD" + m.group(2), s)

def fix_comma_after_od(s):
    return re.sub(r'(OD[\d\.]+)([A-Za-z#])',
                  lambda m: m.group(1) + "," + m.group(2), s)

def fix_space_in_od(s):
    return re.sub(r'\bOD\s+(\d)', r'OD\1', s)

def fix_whitespace(s):
    return re.sub(r'\s+,', ',', s).strip()

def post_correct(desc):
    original = desc
    changes = []
    def apply(fn, label):
        nonlocal desc
        s = fn(desc)
        if s != desc:
            changes.append(label)
            desc = s
    apply(fix_fullwidth, "full-width chars")
    apply(fix_comma_before_color, "comma before color")
    apply(fix_comma_before_od, "comma before OD")
    apply(fix_comma_after_od, "comma after OD")
    apply(fix_space_in_od, "space in OD")
    apply(fix_whitespace, "whitespace")
    return desc, "; ".join(changes) if changes else ""

def preprocess(text):
    text = str(text).strip().upper()
    return text.replace("\uff0c",",").replace("\uff08","(").replace("\uff09",")")

def main(good_path, nogood_path, output_path):

    import os
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print("Loading data...")
    good_df = pd.read_excel(good_path,   sheet_name="Raw")
    nogood_df = pd.read_excel(nogood_path, sheet_name="Raw")

    good_df.columns = good_df.columns.str.strip()
    nogood_df.columns = nogood_df.columns.str.strip()

    good_descs = good_df["Full Item Description"].astype(str).tolist()
    nogood_descs = nogood_df["Full Item Description"].astype(str).tolist()
    good_parts = good_df["Part Number"].astype(str).tolist()
    print(f"- Good: {len(good_descs)}\n- No_Good: {len(nogood_descs)}")

    # TF-IDF
    print("\nTraining TF-IDF model...")
    vectorizer = TfidfVectorizer(analyzer="char_wb", ngram_range=(2,4),
                                 lowercase=True, max_features=8000)
    good_matrix = vectorizer.fit_transform([preprocess(d) for d in good_descs])
    nogood_matrix = vectorizer.transform([preprocess(d) for d in nogood_descs])
    print(f"- Vocabulary: {len(vectorizer.vocabulary_)} features")

    # Match + threshold + post-correct
    print("\nMatching...")
    BATCH = 500
    results = []

    for start in range(0, len(nogood_descs), BATCH):
        end = min(start + BATCH, len(nogood_descs))
        sim_matrix = cosine_similarity(nogood_matrix[start:end], good_matrix)

        for i, sims in enumerate(sim_matrix):
            best_idx = int(np.argmax(sims))
            best_score = float(sims[best_idx])
            original = nogood_descs[start + i]
            part_no = nogood_df["Part Number"].iloc[start + i]

            if best_score >= SIMILARITY_THRESHOLD:
                ml_match = good_descs[best_idx]
                final, corrections = post_correct(ml_match)
                matched_good_pn = good_parts[best_idx]
                status = "AUTO"
            else:
                # score too low -> don't guess, leave blank
                ml_match = good_descs[best_idx]   # kept for reference
                final = ""                      # blank output
                corrections = ""
                matched_good_pn = good_parts[best_idx]
                status = "MANUAL REVIEW NEEDED"

            results.append({
                "No_Good Part#": part_no,
                "Original Description": original,
                "Standardized Description": final,
                "Status": status,
                "Similarity Score": round(best_score, 4),
                "Matched Good Part#": matched_good_pn,
                "ML Match (reference)": ml_match,
                "Post-corrections": corrections if corrections else "none",
            })

        print(f"{end}/{len(nogood_descs)}...", end="\r")

    print("Done.")

    # Write Excel
    print("\nWriting Excel...")
    out_df = pd.DataFrame(results)
    out_df.to_excel(output_path, index=False, sheet_name="Results")

    wb = load_workbook(output_path)
    ws = wb["Results"]

    # Header
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row colors
    # green: auto / red: manual review needed
    green = PatternFill("solid", fgColor="E2EFDA")
    red   = PatternFill("solid", fgColor="F4CCCC")
    for row in ws.iter_rows(min_row=2):
        fill = green if row[3].value == "AUTO" else red
        for cell in row:
            cell.fill = fill

    for i, w in enumerate([15, 65, 65, 22, 14, 18, 65, 35], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(output_path)

    # Summary
    auto = (out_df["Status"] == "AUTO").sum()
    manual = (out_df["Status"] == "MANUAL REVIEW NEEDED").sum()
    total = len(out_df)

    print(f"\nTotal processed: {total}")
    print(f"Auto standardized: {auto}  ({auto/total*100:.1f}%)")
    print(f"Manual review: {manual}  ({manual/total*100:.1f}%)")
    print(f"Threshold used: {SIMILARITY_THRESHOLD}")
    print(f"\nOutput: {output_path}")


if __name__ == "__main__":
    good_path = sys.argv[1] if len(sys.argv) > 1 else "data/1.Good.xlsx"
    nogood_path = sys.argv[2] if len(sys.argv) > 2 else "data/2.No_Good.xlsx"
    output_path = sys.argv[3] if len(sys.argv) > 3 else "outputs/ML_Standardized_Output.xlsx"
    main(good_path, nogood_path, output_path)