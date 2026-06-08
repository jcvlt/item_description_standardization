"""
bm25_rf_standardize.py  —  BM25 features + Random Forest classifier
=====================================================================
Output sheets:
  1. Results                  — before/after for every No_Good description
  2. AI-Learned Standard Rules — standard format, valid values, error classes
"""

import re, sys, os
import pandas as pd
import numpy as np
from rank_bm25 import BM25Okapi
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report
from sklearn.preprocessing import LabelEncoder
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONFIDENCE_THRESHOLD = 0.60
VALIDATION_SIZE      = 0.20
TOP_N                = 30

COLOR_TOKENS = [
    "GRN/YEL-ST","GRN/YEL-RI","WHT/GRY-ST","GN/YW-ST",
    "ORG/WHT","ORG/BLK","RED/WHT","RED/BLK","WHT/BLU","BLU/WHT",
    "WHT/BLK","WHT/RED","WHT/GRN","WHT/GRY","WHT/BRN",
    "YEL/GRN","YEL/WHT","YEL/RED","GRN/YEL","PINK/WHT",
    "LT-BLU","LT-GRN","DK-BLU","DK-GRN","D-BLU","D-GRN",
    "BEIGE","SLATE","DARK","PINK",
    "BLK","RED","WHT","BLU","GRN","YEL","ORG","BRN","GRY","PUR","PNK","VIO",
]
COLOR_PAT = "|".join(re.escape(c) for c in COLOR_TOKENS)

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Arial")
SEC_FILL  = PatternFill("solid", fgColor="D6E4F0")
SEC_FONT  = Font(bold=True, size=11, name="Arial")
BODY_FONT = Font(size=10, name="Arial")
BOLD_FONT = Font(bold=True, size=10, name="Arial")
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN      = Side(style="thin", color="CCCCCC")
BORDER    = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

def hdr(cell, text):
    cell.value = text; cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = CENTER; cell.border = BORDER

def sec(ws, label, title, cols="BC"):
    ws.append([label, title, ""])
    r = ws.max_row
    ws.merge_cells(f"{cols[0]}{r}:{cols[1]}{r}")
    ws[f"A{r}"].value = label; ws[f"A{r}"].fill = SEC_FILL
    ws[f"A{r}"].font = SEC_FONT; ws[f"A{r}"].alignment = LEFT
    ws[f"A{r}"].border = BORDER
    ws[f"B{r}"].value = title; ws[f"B{r}"].fill = SEC_FILL
    ws[f"B{r}"].font = SEC_FONT; ws[f"B{r}"].alignment = LEFT
    ws[f"B{r}"].border = BORDER
    ws[f"C{r}"].fill = SEC_FILL; ws[f"C{r}"].border = BORDER

def body(ws, vals, alt=False):
    ws.append(vals)
    r = ws.max_row
    fill = PatternFill("solid", fgColor="F5F9FF") if alt else None
    for i, col in enumerate("ABCDE"[:len(vals)]):
        cell = ws[f"{col}{r}"]
        cell.font = BODY_FONT; cell.border = BORDER
        cell.alignment = CENTER if i == 0 else LEFT
        if fill: cell.fill = fill
    return r

# ── Tokenizer ─────────────────────────────────────────────────────────────────
def tokenize(text):
    text = str(text).upper().strip()
    text = (text.replace("\uff0c", ",").replace("\uff08", "(")
                .replace("\uff09", ")").replace("\u3000", " "))
    return [t for t in re.split(r"[,\s#/()\-]+", text) if t]

# ── BM25 features ─────────────────────────────────────────────────────────────
def extract_bm25_features(descriptions, bm25_model, n=TOP_N):
    features = []
    for i, desc in enumerate(descriptions):
        scores = bm25_model.get_scores(tokenize(desc))
        top = np.sort(scores)[::-1][:n]
        if len(top) < n:
            top = np.pad(top, (0, n - len(top)))
        features.append(top)
        if (i + 1) % 5000 == 0:
            print(f"    {i+1}/{len(descriptions)}...", end="\r")
    return np.array(features)

# ── Hand-crafted features ─────────────────────────────────────────────────────
def hand_crafted_features(descriptions):
    features = []
    for d in descriptions:
        d = str(d)
        features.append([
            int(bool(re.match(r"^CBL\.WIRE", d))),
            int(bool(re.search(r"WIRE:", d))),
            int(",," in d),
            int(bool(re.search(r"\)(" + COLOR_PAT + r")([,\s]|$)", d))),
            int(bool(re.match(r"^\s+", d))),
            int(bool(re.search(r"\s+,", d))),
            int(bool(re.search(r"[^\x00-\x7F]", d))),
            d.count(","), len(d),
        ])
    return np.array(features, dtype=float)

# ── Error injection ───────────────────────────────────────────────────────────
def inject_error(desc, error_type):
    d = desc.strip()
    if error_type == "clean": return d
    elif error_type == "dot_separator":
        return re.sub(r"^CBL,WIRE", "CBL.WIRE", d)
    elif error_type == "colon_separator":
        return re.sub(r"WIRE,", "WIRE:", d, count=1)
    elif error_type == "double_comma":
        idx = d.find(",")
        return d[:idx] + ",," + d[idx+1:] if idx != -1 else d
    elif error_type == "missing_comma_before_color":
        return re.sub(r",(" + COLOR_PAT + r")([,\s]|$)",
                      lambda m: m.group(1) + m.group(2), d)
    elif error_type == "leading_space": return "  " + d
    elif error_type == "space_before_comma":
        idx = d.find(",")
        return d[:idx] + " ," + d[idx+1:] if idx != -1 else d
    return d

def generate_synthetic_data(good_descs):
    error_types = ["clean","dot_separator","colon_separator","double_comma",
                   "missing_comma_before_color","leading_space","space_before_comma"]
    descriptions, labels = [], []
    for desc in good_descs:
        for et in error_types:
            if et == "missing_comma_before_color":
                if not re.search(r",(" + COLOR_PAT + r")[,\s]", desc):
                    continue
            descriptions.append(inject_error(desc, et))
            labels.append(et)
    return descriptions, labels

# ── Correction ────────────────────────────────────────────────────────────────
def apply_correction(desc, predicted_error):
    d = str(desc).strip(); original = d
    if predicted_error == "clean": return d, "none"
    elif predicted_error == "dot_separator":
        d = re.sub(r"^CBL\.WIRE", "CBL,WIRE", d)
    elif predicted_error == "colon_separator":
        d = re.sub(r"WIRE:", "WIRE,", d)
    elif predicted_error == "double_comma":
        d = re.sub(r",{2,}", ",", d)
    elif predicted_error == "missing_comma_before_color":
        d = re.sub(r"\)(" + COLOR_PAT + r")([,\s]|$)",
                   lambda m: ")," + m.group(1) + m.group(2), d)
    elif predicted_error == "leading_space": d = d.lstrip()
    elif predicted_error == "space_before_comma":
        d = re.sub(r"\s+,", ",", d)
    return d, (predicted_error if d != original else "none")

# ── Build Sheet 2: AI-Learned Standard Rules ──────────────────────────────────
def build_rules_sheet(wb, good_descs, val_report_dict, train_acc, val_acc):
    ws = wb.create_sheet("AI-Learned Standard Rules")

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"].value = "AI-Learned Standard Description Rules"
    ws["A1"].font  = Font(bold=True, size=14, color="1F4E79", name="Arial")
    ws["A1"].alignment = CENTER
    ws["A1"].fill = PatternFill("solid", fgColor="EBF3FB")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    ws["A2"].value = f"Learned from {len(good_descs):,} Good descriptions  |  Model: BM25 + Random Forest  |  Validation accuracy: {val_acc*100:.2f}%"
    ws["A2"].font  = Font(italic=True, size=10, color="555555", name="Arial")
    ws["A2"].alignment = CENTER
    ws.append([])

    # ── Section 1: Standard Format ────────────────────────────────────────────
    sec(ws, "SECTION 1", "Standard Description Format")

    ws.append(["", "Full format:", "CBL,WIRE,[Standard],[Gauge#(Strands/Diameter+Material)],[Color],[OD],[Vendor Part#]"])
    r = ws.max_row
    ws[f"B{r}"].font = BOLD_FONT; ws[f"B{r}"].border = BORDER; ws[f"B{r}"].alignment = LEFT
    ws[f"C{r}"].font = Font(bold=True, size=10, color="1F4E79", name="Arial")
    ws[f"C{r}"].border = BORDER; ws[f"C{r}"].alignment = LEFT

    ws.append(["", "Example:", "CBL,WIRE,MIL-W-16878E/1,24#(19/36TA),BLK,OD1.14,ALPHA#1854/19-2"])
    r = ws.max_row
    ws[f"B{r}"].font = BOLD_FONT; ws[f"B{r}"].border = BORDER; ws[f"B{r}"].alignment = LEFT
    ws[f"C{r}"].font = BODY_FONT; ws[f"C{r}"].border = BORDER; ws[f"C{r}"].alignment = LEFT

    ws.append([])

    for field, example, desc in [
        ("Field 1 — CBL,WIRE",    "Fixed prefix",                        "Always present, always uppercase, never changes"),
        ("Field 2 — [Standard]",  "e.g. UL1007, SAE J1128-TXL",         "Wire/cable standard specification"),
        ("Field 3 — [Gauge]",     "e.g. 22#(19/34TA)",                   "Gauge number, strand count, wire diameter, surface treatment (TA=tin coated, BA=bare)"),
        ("Field 4 — [Color]",     "e.g. BLK, RED, GRN/YEL",             "Standard color abbreviation — see Section 3 for valid codes"),
        ("Field 5 — [OD]",        "e.g. OD1.32, OD2.0",                  "Outer diameter — must be prefixed with OD, no space (OD1.32 not OD 1.32)"),
        ("Field 6 — [Vendor]",    "e.g. ALPHA#1855/19-2",                "Manufacturer part number"),
    ]:
        ws.append(["", field, example, desc])
        r = ws.max_row
        ws[f"A{r}"].border = BORDER
        ws[f"B{r}"].font = BOLD_FONT; ws[f"B{r}"].border = BORDER; ws[f"B{r}"].alignment = LEFT
        ws[f"C{r}"].font = BODY_FONT; ws[f"C{r}"].border = BORDER; ws[f"C{r}"].alignment = CENTER
        ws[f"D{r}"].font = BODY_FONT; ws[f"D{r}"].border = BORDER; ws[f"D{r}"].alignment = LEFT

    ws.append([])

    # ── Section 2: Formatting Rules ───────────────────────────────────────────
    sec(ws, "SECTION 2", "Formatting Rules")

    for rule, wrong, correct in [
        ("Separator between fields",          "CBL,WIRE,22#,BLK ,OD1.32",      "CBL,WIRE,22#,BLK,OD1.32  (comma, no spaces)"),
        ("Comma before color code",           "(19/36TA)BLK",                   "(19/36TA),BLK"),
        ("Comma before OD",                   "BLK OD1.32",                     "BLK,OD1.32"),
        ("Comma after OD",                    "OD1.32ALPHA",                    "OD1.32,ALPHA"),
        ("No space inside OD",               "OD 1.32",                        "OD1.32"),
        ("Full-width punctuation → ASCII",    "CBL，WIRE（19/36TA）",            "CBL,WIRE,(19/36TA)"),
        ("No double comma",                   "CBL,,WIRE",                      "CBL,WIRE"),
        ("No dot separator",                  "CBL.WIRE",                       "CBL,WIRE"),
        ("No colon separator",                "CBL,WIRE:MIL",                   "CBL,WIRE,MIL"),
        ("No leading spaces",                 "  CBL,WIRE,...",                 "CBL,WIRE,..."),
    ]:
        ws.append(["", rule, wrong, correct])
        r = ws.max_row
        ws[f"A{r}"].border = BORDER
        ws[f"B{r}"].font = BODY_FONT; ws[f"B{r}"].border = BORDER; ws[f"B{r}"].alignment = LEFT
        ws[f"C{r}"].font = Font(size=10, color="C00000", name="Arial")
        ws[f"C{r}"].border = BORDER; ws[f"C{r}"].alignment = LEFT
        ws[f"D{r}"].font = Font(size=10, color="375623", name="Arial")
        ws[f"D{r}"].border = BORDER; ws[f"D{r}"].alignment = LEFT

    ws.append([])

    # ── Section 3: Valid Color Codes ──────────────────────────────────────────
    sec(ws, "SECTION 3", "Valid Color Codes (learned from Good data)")

    color_pat = r'\b(' + COLOR_PAT + r')\b'
    colors = []
    for d in good_descs:
        colors.extend(re.findall(color_pat, d))
    color_counts = Counter(colors).most_common()

    # Display in rows of 6
    row_vals = []
    for color, cnt in color_counts:
        row_vals.append(f"{color} ({cnt})")
        if len(row_vals) == 6:
            ws.append([""] + row_vals)
            r = ws.max_row
            ws[f"A{r}"].border = BORDER
            for c in "BCDEFG"[:6]:
                ws[f"{c}{r}"].font = BODY_FONT
                ws[f"{c}{r}"].alignment = CENTER
                ws[f"{c}{r}"].border = BORDER
            row_vals = []
    if row_vals:
        ws.append([""] + row_vals)
        r = ws.max_row
        ws[f"A{r}"].border = BORDER
        for c in "BCDEFG"[:len(row_vals)]:
            ws[f"{c}{r}"].font = BODY_FONT
            ws[f"{c}{r}"].alignment = CENTER
            ws[f"{c}{r}"].border = BORDER

    ws.append([])

    # ── Section 4: Top Wire Standards ────────────────────────────────────────
    sec(ws, "SECTION 4", "Top Wire Standards (learned from Good data)")

    standards = []
    for d in good_descs:
        parts = d.split(",")
        if len(parts) >= 3:
            standards.append(parts[2].strip())

    std_vals = []
    for std, cnt in Counter(standards).most_common(18):
        std_vals.append(f"{std} ({cnt})")
        if len(std_vals) == 3:
            ws.append([""] + std_vals)
            r = ws.max_row
            ws[f"A{r}"].border = BORDER
            for c in "BCD":
                ws[f"{c}{r}"].font = BODY_FONT
                ws[f"{c}{r}"].alignment = LEFT
                ws[f"{c}{r}"].border = BORDER
            std_vals = []
    if std_vals:
        ws.append([""] + std_vals)
        r = ws.max_row
        ws[f"A{r}"].border = BORDER
        for c in "BCD"[:len(std_vals)]:
            ws[f"{c}{r}"].font = BODY_FONT
            ws[f"{c}{r}"].alignment = LEFT
            ws[f"{c}{r}"].border = BORDER

    ws.append([])

    # ── Section 5: ML Error Classes ───────────────────────────────────────────
    sec(ws, "SECTION 5", "Error Classes Learned by the Model (BM25 + Random Forest)")

    ws.append(["", "Error Class", "Wrong → Correct", "F1-Score"])
    r = ws.max_row
    for c in "ABCD":
        ws[f"{c}{r}"].font = BOLD_FONT
        ws[f"{c}{r}"].border = BORDER
        ws[f"{c}{r}"].alignment = CENTER
        ws[f"{c}{r}"].fill = PatternFill("solid", fgColor="EBF3FB")

    error_classes = [
        ("clean",                     "(already correct) → (no change)"),
        ("dot_separator",             "CBL.WIRE → CBL,WIRE"),
        ("colon_separator",           "CBL,WIRE:MIL → CBL,WIRE,MIL"),
        ("double_comma",              "CBL,,WIRE → CBL,WIRE"),
        ("missing_comma_before_color","(19/36TA)BLK → (19/36TA),BLK"),
        ("leading_space",             "  CBL,WIRE → CBL,WIRE"),
        ("space_before_comma",        "CBL ,WIRE → CBL,WIRE"),
    ]
    row_colors = ["F4CCCC","FFF2CC","FCE4D6","E2EFDA","D9EAD3","CFE2F3","EAD1DC"]
    for i, (cls, example) in enumerate(error_classes):
        f1 = round(val_report_dict[cls]["f1-score"], 3) if cls in val_report_dict and isinstance(val_report_dict[cls], dict) else "—"
        ws.append(["", cls, example, f1])
        r = ws.max_row
        fill = PatternFill("solid", fgColor=row_colors[i])
        for c in "ABCD":
            ws[f"{c}{r}"].fill = fill
            ws[f"{c}{r}"].font = BODY_FONT
            ws[f"{c}{r}"].border = BORDER
            ws[f"{c}{r}"].alignment = CENTER if c in "AD" else LEFT

    ws.append([])
    note_r = ws.max_row
    ws[f"A{note_r}"].value = f"Model validation accuracy: {val_acc*100:.2f}%  |  Training accuracy: {train_acc*100:.2f}%  |  Confidence threshold: {CONFIDENCE_THRESHOLD}"
    ws[f"A{note_r}"].font = Font(italic=True, size=9, color="888888", name="Arial")
    ws.merge_cells(f"A{note_r}:D{note_r}")

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18


# ── Main ──────────────────────────────────────────────────────────────────────
def main(good_path, nogood_path, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print("Loading data...")
    good_df   = pd.read_excel(good_path,   sheet_name="Raw")
    nogood_df = pd.read_excel(nogood_path, sheet_name="Raw")
    good_df.columns   = good_df.columns.str.strip()
    nogood_df.columns = nogood_df.columns.str.strip()
    good_descs   = good_df["Full Item Description"].astype(str).tolist()
    nogood_descs = nogood_df["Full Item Description"].astype(str).tolist()
    print(f"  Good: {len(good_descs)}  |  No Good: {len(nogood_descs)}")

    print("\nBuilding BM25 index...")
    tokenized_good = [tokenize(d) for d in good_descs]
    bm25 = BM25Okapi(tokenized_good)

    print("\nGenerating synthetic training data...")
    all_descs, all_labels = generate_synthetic_data(good_descs)
    X_train_raw, X_val_raw, y_train, y_val = train_test_split(
        all_descs, all_labels, test_size=VALIDATION_SIZE,
        random_state=42, stratify=all_labels)
    print(f"  Train: {len(X_train_raw)}  |  Validation: {len(X_val_raw)}")

    print("\nExtracting BM25 features...")
    X_train = np.hstack([extract_bm25_features(X_train_raw, bm25), hand_crafted_features(X_train_raw)])
    X_val   = np.hstack([extract_bm25_features(X_val_raw,   bm25), hand_crafted_features(X_val_raw)])

    le = LabelEncoder()
    y_train_enc = le.fit_transform(y_train)
    y_val_enc   = le.transform(y_val)

    print("\nTraining Random Forest (200 trees)...")
    clf = RandomForestClassifier(n_estimators=200, max_depth=20,
                                  min_samples_leaf=2, n_jobs=-1, random_state=42)
    clf.fit(X_train, y_train_enc)

    train_acc = clf.score(X_train, y_train_enc)
    val_acc   = clf.score(X_val,   y_val_enc)
    print(f"  Training accuracy  : {train_acc*100:.2f}%")
    print(f"  Validation accuracy: {val_acc*100:.2f}%")

    y_val_pred = clf.predict(X_val)
    val_report_dict = classification_report(
        y_val_enc, y_val_pred, target_names=le.classes_, digits=3, output_dict=True)
    print("\n  Per-class report:")
    print(classification_report(y_val_enc, y_val_pred, target_names=le.classes_, digits=3))

    print("Predicting on No_Good descriptions...")
    X_test = np.hstack([extract_bm25_features(nogood_descs, bm25),
                        hand_crafted_features(nogood_descs)])
    predicted_labels  = le.inverse_transform(clf.predict(X_test))
    confidence_scores = clf.predict_proba(X_test).max(axis=1)

    good_set = set(d.strip() for d in good_descs)
    results, counts = [], {"AUTO": 0, "MANUAL REVIEW NEEDED": 0}
    for i, original in enumerate(nogood_descs):
        part_no    = nogood_df["Part Number"].iloc[i]
        pred_error = predicted_labels[i]
        confidence = float(confidence_scores[i])
        if original.strip() in good_set:
            final, correction, status = original.strip(), "none", "AUTO"
            confidence, pred_error = 1.0, "clean"
        elif confidence >= CONFIDENCE_THRESHOLD:
            final, correction = apply_correction(original, pred_error)
            status = "AUTO"
        else:
            final, correction, status = "", "", "MANUAL REVIEW NEEDED"
        counts[status] += 1
        results.append({
            "No_Good Part#"           : part_no,
            "Original Description"    : original,
            "Standardized Description": final,
            "Status"                  : status,
            "Predicted Error Type"    : pred_error,
            "Confidence"              : round(confidence, 4),
            "Correction Applied"      : correction,
        })

    print("\nWriting Excel...")
    out_df = pd.DataFrame(results)
    out_df.to_excel(output_path, index=False, sheet_name="Results")

    wb = load_workbook(output_path)

    # Style Results sheet
    ws_r = wb["Results"]
    for cell in ws_r[1]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER
    green = PatternFill("solid", fgColor="E2EFDA")
    red   = PatternFill("solid", fgColor="F4CCCC")
    for row in ws_r.iter_rows(min_row=2):
        fill = green if row[3].value == "AUTO" else red
        for cell in row: cell.fill = fill
    for idx, w in enumerate([15, 65, 65, 22, 26, 12, 26], 1):
        ws_r.column_dimensions[get_column_letter(idx)].width = w
    ws_r.freeze_panes = "A2"

    # Build Sheet 2
    build_rules_sheet(wb, good_descs, val_report_dict, train_acc, val_acc)

    wb.save(output_path)

    total = len(out_df); auto = counts["AUTO"]; manual = counts["MANUAL REVIEW NEEDED"]
    print(f"\n{'='*55}")
    print(f"  Total processed      : {total}")
    print(f"  Auto standardized    : {auto}  ({auto/total*100:.1f}%)")
    print(f"  Manual review        : {manual}  ({manual/total*100:.1f}%)")
    print(f"\n  Model (BM25 + Random Forest):")
    print(f"    Training accuracy  : {train_acc*100:.2f}%")
    print(f"    Validation accuracy: {val_acc*100:.2f}%")
    print(f"\n  Output sheets:")
    print(f"    1. Results                 — before/after (Output 1)")
    print(f"    2. AI-Learned Standard Rules — rules + values + error classes (Output 2)")
    print(f"\n  Output: {output_path}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    good_path   = sys.argv[1] if len(sys.argv) > 1 else "data/1.Good.xlsx"
    nogood_path = sys.argv[2] if len(sys.argv) > 2 else "data/2.No_Good.xlsx"
    output_path = sys.argv[3] if len(sys.argv) > 3 else "outputs/BM25_RF_Standardized_Output.xlsx"
    main(good_path, nogood_path, output_path)
