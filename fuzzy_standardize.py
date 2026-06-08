"""
Item Description Standardization using Fuzzy Matching (Edit Distance)
======================================================================
Uses Python's built-in difflib.SequenceMatcher to find the most similar
Good description for each No_Good entry.

How it works:
  1. Load Good.xlsx as the reference set
  2. For each No_Good description, compare it against all Good descriptions
     using SequenceMatcher (ratio = similarity score between 0 and 1)
  3. If best score >= 0.95, use that Good description as the standardized output
  4. If best score < 0.95, leave blank and flag for manual review
  5. Apply post-correction rules to the matched output

Difference from TF-IDF:
  - TF-IDF:  vectorizes text into numbers, compares vector angles
  - Fuzzy:   directly counts how many characters need to change (edit distance)

Usage:
  python fuzzy_standardize.py                        # default paths
  python fuzzy_standardize.py good.xlsx nogood.xlsx output.xlsx
"""

import re, sys
import pandas as pd
from difflib import SequenceMatcher, get_close_matches
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SIMILARITY_THRESHOLD = 0.95

COLOR_TOKENS = [
    "GRN/YEL-ST","GRN/YEL-RI","WHT/GRY-ST","GN/YW-ST",
    "ORG/WHT","ORG/BLK","RED/WHT","RED/BLK","WHT/BLU","BLU/WHT",
    "WHT/BLK","WHT/RED","WHT/GRN","WHT/GRY","WHT/BRN",
    "YEL/GRN","YEL/WHT","YEL/RED","GRN/YEL","PINK/WHT",
    "LT-BLU","LT-GRN","DK-BLU","DK-GRN","D-BLU","D-GRN",
    "BEIGE","SLATE","DARK","PINK",
    "BLK","RED","WHT","BLU","GRN","YEL","ORG","BRN","GRY","PUR","PNK","VIO",
]
COLOR_PAT = "|".join(re.escape(c) for c in COLOR_TOKENS)

def fix_fullwidth(s):
    return (s.replace("\uff0c",",").replace("\uff08","(")
             .replace("\uff09",")").replace("\u3000"," "))

def fix_comma_before_color(s):
    return re.sub(r'\)(' + COLOR_PAT + r')([,\s]|$)',
                  lambda m: ")," + m.group(1) + m.group(2), s)

def fix_comma_before_od(s):
    return re.sub(r'(?<!\))([A-Za-z])OD(\d)',
                  lambda m: m.group(1) + ",OD" + m.group(2), s)

def fix_comma_after_od(s):
    return re.sub(r'(OD[\d\.]+)([A-Za-z#])',
                  lambda m: m.group(1) + "," + m.group(2), s)

def fix_space_in_od(s):
    return re.sub(r'\bOD\s+(\d)', r'OD\1', s)

def fix_whitespace(s):
    return re.sub(r'\s+,', ',', s).strip()

def post_correct(desc):
    original = desc
    changes = []
    def apply(fn, label):
        nonlocal desc
        s = fn(desc)
        if s != desc:
            changes.append(label)
            desc = s
    apply(fix_fullwidth,          "full-width chars")
    apply(fix_comma_before_color, "comma before color")
    apply(fix_comma_before_od,    "comma before OD")
    apply(fix_comma_after_od,     "comma after OD")
    apply(fix_space_in_od,        "space in OD")
    apply(fix_whitespace,         "whitespace")
    return desc, "; ".join(changes) if changes else ""

def preprocess(text):
    return str(text).strip().upper()

def main(good_path, nogood_path, output_path):

    import os
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print("Loading data...")
    good_df   = pd.read_excel(good_path,   sheet_name="Raw")
    nogood_df = pd.read_excel(nogood_path, sheet_name="Raw")
    good_df.columns   = good_df.columns.str.strip()
    nogood_df.columns = nogood_df.columns.str.strip()

    good_descs   = good_df["Full Item Description"].astype(str).tolist()
    nogood_descs = nogood_df["Full Item Description"].astype(str).tolist()
    good_parts   = good_df["Part Number"].astype(str).tolist()

    good_processed   = [preprocess(d) for d in good_descs]
    nogood_processed = [preprocess(d) for d in nogood_descs]

    # Build lookup dict for instant exact matches
    good_lookup = {g: i for i, g in enumerate(good_processed)}

    print(f"  Good: {len(good_descs)}   No_Good: {len(nogood_descs)}")
    print("\nFuzzy matching...")

    results = []

    for i, (nogood_raw, nogood_pre) in enumerate(zip(nogood_descs, nogood_processed)):

        part_no = nogood_df["Part Number"].iloc[i]

        # Fast path: exact match
        if nogood_pre in good_lookup:
            best_idx   = good_lookup[nogood_pre]
            best_score = 1.0
        else:
            # Use get_close_matches for fast candidate filtering
            # then score only top candidates
            candidates = get_close_matches(nogood_pre, good_processed, n=5, cutoff=0.6)

            if candidates:
                best_score = -1
                best_idx   = 0
                for cand in candidates:
                    idx   = good_lookup[cand]
                    score = SequenceMatcher(None, nogood_pre, cand).ratio()
                    if score > best_score:
                        best_score = score
                        best_idx   = idx
            else:
                # No candidate found above 0.6 cutoff
                best_score = 0.0
                best_idx   = 0

        if best_score >= SIMILARITY_THRESHOLD:
            fuzzy_match        = good_descs[best_idx]
            final, corrections = post_correct(fuzzy_match)
            matched_good_pn    = good_parts[best_idx]
            status             = "AUTO"
        else:
            fuzzy_match        = good_descs[best_idx] if best_idx else ""
            final              = ""
            corrections        = ""
            matched_good_pn    = good_parts[best_idx] if best_idx else ""
            status             = "MANUAL REVIEW NEEDED"

        results.append({
            "No_Good Part#":              part_no,
            "Original Description":       nogood_raw,
            "Standardized Description":   final,
            "Status":                     status,
            "Similarity Score":           round(best_score, 4),
            "Matched Good Part#":         matched_good_pn,
            "Fuzzy Match (reference)":    fuzzy_match,
            "Post-corrections":           corrections if corrections else "none",
        })

        if (i + 1) % 500 == 0:
            print(f"  {i+1}/{len(nogood_descs)}...", end="\r")

    print(f"  {len(nogood_descs)}/{len(nogood_descs)}... Done!     ")

    print("\nWriting Excel...")
    out_df = pd.DataFrame(results)
    out_df.to_excel(output_path, index=False, sheet_name="Results")

    wb = load_workbook(output_path)
    ws = wb["Results"]

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    green = PatternFill("solid", fgColor="E2EFDA")
    red   = PatternFill("solid", fgColor="F4CCCC")
    for row in ws.iter_rows(min_row=2):
        fill = green if row[3].value == "AUTO" else red
        for cell in row:
            cell.fill = fill

    for i, w in enumerate([15, 65, 65, 22, 14, 18, 65, 35], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(output_path)

    auto   = (out_df["Status"] == "AUTO").sum()
    manual = (out_df["Status"] == "MANUAL REVIEW NEEDED").sum()
    total  = len(out_df)

    print(f"\n{'='*55}")
    print(f"  Total processed       : {total}")
    print(f"  ✅ Auto standardized  : {auto}  ({auto/total*100:.1f}%)")
    print(f"  ⚠️  Manual review     : {manual}  ({manual/total*100:.1f}%)")
    print(f"  Threshold used        : {SIMILARITY_THRESHOLD}")
    print(f"\n  Output: {output_path}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    good_path   = sys.argv[1] if len(sys.argv) > 1 else "data/1.Good.xlsx"
    nogood_path = sys.argv[2] if len(sys.argv) > 2 else "data/2.No_Good.xlsx"
    output_path = sys.argv[3] if len(sys.argv) > 3 else "outputs/Fuzzy_Standardized_Output.xlsx"
    main(good_path, nogood_path, output_path)