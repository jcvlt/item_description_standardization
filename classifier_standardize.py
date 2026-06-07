"""
classifier_standardize.py  —  Stage ML: Error classifier + targeted correction

HOW IT WORKS:
  1. SYNTHETIC DATA GENERATION
     Takes the 4,511 Good descriptions and injects known error patterns
     artificially to create labeled training pairs:
       (broken_description, error_label)

  2. TRAIN / VALIDATION SPLIT  (80% / 20%, stratified)
     Keeps 20% of synthetic data unseen during training to measure
     how well the model generalizes — not just memorizes.

  3. FEATURE EXTRACTION
     - Character n-gram TF-IDF (captures local character patterns)
     - 9 hand-crafted binary/numeric flags (direct error signals)

  4. TRAINING — Random Forest (200 trees)
     Learns which feature combinations predict which error type.
     Reports both training AND validation accuracy.

  5. PREDICTION + CORRECTION
     For each No Good description:
       a. Classifier predicts the most likely error type
       b. Targeted fix applied for that specific error type
       c. Confidence < threshold → MANUAL REVIEW NEEDED

ERROR CLASSES:
  clean                      → already correct
  dot_separator              → CBL.WIRE  →  CBL,WIRE
  colon_separator            → WIRE:     →  WIRE,
  double_comma               → ,,        →  ,
  missing_comma_before_color → )BLK      →  ),BLK
  leading_space              →  CBL,WIRE →  CBL,WIRE
  space_before_comma         → CBL ,WIRE →  CBL,WIRE
"""
import os
import re, sys
import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report
from sklearn.preprocessing import LabelEncoder
from scipy.sparse import hstack, csr_matrix
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

CONFIDENCE_THRESHOLD = 0.60
VALIDATION_SIZE      = 0.20   # 20% of synthetic data held out for validation

# COLOR TOKENS 
COLOR_TOKENS = [
    "GRN/YEL-ST","GRN/YEL-RI","WHT/GRY-ST","GN/YW-ST",
    "ORG/WHT","ORG/BLK","RED/WHT","RED/BLK","WHT/BLU","BLU/WHT",
    "WHT/BLK","WHT/RED","WHT/GRN","WHT/GRY","WHT/BRN",
    "YEL/GRN","YEL/WHT","YEL/RED","GRN/YEL","PINK/WHT",
    "LT-BLU","LT-GRN","DK-BLU","DK-GRN","D-BLU","D-GRN",
    "BEIGE","SLATE","DARK","PINK",
    "BLK","RED","WHT","BLU","GRN","YEL","ORG","BRN","GRY","PUR","PNK","VIO",
]
COLOR_PAT = "|".join(re.escape(c) for c in COLOR_TOKENS)


# STEP 1
def inject_error(desc: str, error_type: str) -> str:
    d = desc.strip()
    if error_type == "clean":
        return d
    elif error_type == "dot_separator":
        return re.sub(r"^CBL,WIRE", "CBL.WIRE", d)
    elif error_type == "colon_separator":
        return re.sub(r"WIRE,", "WIRE:", d, count=1)
    elif error_type == "double_comma":
        idx = d.find(",")
        return d[:idx] + ",," + d[idx+1:] if idx != -1 else d
    elif error_type == "missing_comma_before_color":
        return re.sub(r",(" + COLOR_PAT + r")([,\s]|$)",
                      lambda m: m.group(1) + m.group(2), d)
    elif error_type == "leading_space":
        return "  " + d
    elif error_type == "space_before_comma":
        idx = d.find(",")
        return d[:idx] + " ," + d[idx+1:] if idx != -1 else d
    return d


def generate_synthetic_data(good_descs: list) -> tuple[list, list]:
    error_types = [
        "clean", "dot_separator", "colon_separator", "double_comma",
        "missing_comma_before_color", "leading_space", "space_before_comma",
    ]
    descriptions, labels = [], []
    for desc in good_descs:
        for et in error_types:
            if et == "missing_comma_before_color":
                if not re.search(r",(" + COLOR_PAT + r")[,\s]", desc):
                    continue
            descriptions.append(inject_error(desc, et))
            labels.append(et)
    return descriptions, labels


# STEP 2
def hand_crafted_features(descriptions: list) -> np.ndarray:
    features = []
    for d in descriptions:
        d = str(d)
        features.append([
            int(bool(re.match(r"^CBL\.WIRE", d))),
            int(bool(re.search(r"WIRE:", d))),
            int(",," in d),
            int(bool(re.search(r"\)(" + COLOR_PAT + r")([,\s]|$)", d))),
            int(bool(re.match(r"^\s+", d))),
            int(bool(re.search(r"\s+,", d))),
            int(bool(re.search(r"[^\x00-\x7F]", d))),
            d.count(","),
            len(d),
        ])
    return np.array(features, dtype=float)


# STEP 3
def apply_correction(desc: str, predicted_error: str) -> tuple[str, str]:
    d = str(desc).strip()
    original = d
    if predicted_error == "clean":
        return d, "none"
    elif predicted_error == "dot_separator":
        d = re.sub(r"^CBL\.WIRE", "CBL,WIRE", d)
    elif predicted_error == "colon_separator":
        d = re.sub(r"WIRE:", "WIRE,", d)
    elif predicted_error == "double_comma":
        d = re.sub(r",{2,}", ",", d)
    elif predicted_error == "missing_comma_before_color":
        d = re.sub(r"\)(" + COLOR_PAT + r")([,\s]|$)",
                   lambda m: ")," + m.group(1) + m.group(2), d)
    elif predicted_error == "leading_space":
        d = d.lstrip()
    elif predicted_error == "space_before_comma":
        d = re.sub(r"\s+,", ",", d)
    return d, (predicted_error if d != original else "none")


# MAIN 
def main(good_path, nogood_path, output_path):

    import os
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print("Loading data...")
    good_df   = pd.read_excel(good_path,   sheet_name="Raw")
    nogood_df = pd.read_excel(nogood_path, sheet_name="Raw")
    good_df.columns   = good_df.columns.str.strip()
    nogood_df.columns = nogood_df.columns.str.strip()

    good_descs   = good_df["Full Item Description"].astype(str).tolist()
    nogood_descs = nogood_df["Full Item Description"].astype(str).tolist()
    good_parts   = good_df["Part Number"].astype(str).tolist()
    print(f"  Good   : {len(good_descs)}")
    print(f"  No Good: {len(nogood_descs)}")

    # ── Generate synthetic data ───────────────────────────────────────────────
    print("\nGenerating synthetic training data...")
    all_descs, all_labels = generate_synthetic_data(good_descs)
    print(f"  Total synthetic examples: {len(all_descs)}")

    # ── Train / Validation split (80/20, stratified)
    X_train_raw, X_val_raw, y_train, y_val = train_test_split(
        all_descs, all_labels,
        test_size=VALIDATION_SIZE,
        random_state=42,
        stratify=all_labels     # ensures all classes represented in both sets
    )
    print(f" Train examples : {len(X_train_raw)}")
    print(f" Validation examples: {len(X_val_raw)}")

    # ── Feature extraction 
    print("\nExtracting features...")
    tfidf = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 4),
                            lowercase=True, max_features=6000)
    X_tfidf_train = tfidf.fit_transform(X_train_raw)
    X_tfidf_val   = tfidf.transform(X_val_raw)

    X_train = hstack([X_tfidf_train, csr_matrix(hand_crafted_features(X_train_raw))])
    X_val   = hstack([X_tfidf_val,   csr_matrix(hand_crafted_features(X_val_raw))])

    le = LabelEncoder()
    y_train_enc = le.fit_transform(y_train)
    y_val_enc   = le.transform(y_val)
    print(f"  Feature matrix shape: {X_train.shape}")
    print(f"  Classes: {list(le.classes_)}")

    # ── Train Random Forest ───────────────────────────────────────────────────
    print("\nTraining Random Forest (200 trees)...")
    clf = RandomForestClassifier(
        n_estimators=200,
        max_depth=20,
        min_samples_leaf=2,
        n_jobs=-1,
        random_state=42,
    )
    clf.fit(X_train, y_train_enc)

    train_acc = clf.score(X_train, y_train_enc)
    val_acc   = clf.score(X_val,   y_val_enc)
    gap       = train_acc - val_acc

    print(f"\n  Training accuracy  : {train_acc*100:.2f}%")
    print(f"  Validation accuracy: {val_acc*100:.2f}%")
    print(f"  Overfit gap        : {gap*100:.2f}%  {'(acceptable)' if gap < 0.05 else '(WARNING: possible overfit)'}")

    print("\n  Per-class report on VALIDATION set:")
    y_val_pred = clf.predict(X_val)
    report = classification_report(y_val_enc, y_val_pred,
                                   target_names=le.classes_, digits=3)
    print(report)

    # ── Predict on No Good descriptions ──────────────────────────────────────
    print("Predicting error types on No Good descriptions...")
    X_tfidf_test = tfidf.transform(nogood_descs)
    X_test = hstack([X_tfidf_test, csr_matrix(hand_crafted_features(nogood_descs))])

    predicted_classes = clf.predict(X_test)
    predicted_probs   = clf.predict_proba(X_test)
    predicted_labels  = le.inverse_transform(predicted_classes)
    confidence_scores = predicted_probs.max(axis=1)

    # ── Apply corrections ─────────────────────────────────────────────────────
    good_set = set(d.strip() for d in good_descs)
    results, counts = [], {"AUTO": 0, "MANUAL REVIEW NEEDED": 0}

    for i, original in enumerate(nogood_descs):
        part_no    = nogood_df["Part Number"].iloc[i]
        pred_error = predicted_labels[i]
        confidence = float(confidence_scores[i])

        if original.strip() in good_set:
            final, correction, status, confidence, pred_error = \
                original.strip(), "none", "AUTO", 1.0, "clean"
        elif confidence >= CONFIDENCE_THRESHOLD:
            final, correction = apply_correction(original, pred_error)
            status = "AUTO"
        else:
            final, correction, status = "", "", "MANUAL REVIEW NEEDED"

        counts[status] += 1
        results.append({
            "No_Good Part#"           : part_no,
            "Original Description"    : original,
            "Standardized Description": final,
            "Status"                  : status,
            "Predicted Error Type"    : pred_error,
            "Confidence"              : round(confidence, 4),
            "Correction Applied"      : correction,
        })

    # ── Write Excel ───────────────────────────────────────────────────────────
    print("\nWriting Excel...")
    out_df = pd.DataFrame(results)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False, sheet_name="Results")

        # Validation metrics sheet
        val_report_dict = classification_report(
            y_val_enc, y_val_pred, target_names=le.classes_,
            digits=3, output_dict=True
        )
        metrics_rows = []
        for label, metrics in val_report_dict.items():
            if isinstance(metrics, dict):
                metrics_rows.append({
                    "Class"    : label,
                    "Precision": round(metrics["precision"], 4),
                    "Recall"   : round(metrics["recall"], 4),
                    "F1-Score" : round(metrics["f1-score"], 4),
                    "Support"  : int(metrics["support"]),
                })
        metrics_df = pd.DataFrame(metrics_rows)
        # Add summary rows
        summary_rows = pd.DataFrame([
            {"Class": "─" * 20},
            {"Class": "Training Accuracy",   "Precision": round(train_acc, 4)},
            {"Class": "Validation Accuracy", "Precision": round(val_acc, 4)},
            {"Class": "Overfit Gap",         "Precision": round(gap, 4)},
            {"Class": "Train examples",      "Support": len(X_train_raw)},
            {"Class": "Validation examples", "Support": len(X_val_raw)},
        ])
        pd.concat([metrics_df, summary_rows], ignore_index=True).to_excel(
            writer, index=False, sheet_name="Validation Metrics"
        )

    # ── Style Excel ───────────────────────────────────────────────────────────
    wb = load_workbook(output_path)

    for sheet_name in ["Results", "Validation Metrics"]:
        ws = wb[sheet_name]
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws = wb["Results"]
    green = PatternFill("solid", fgColor="E2EFDA")
    red   = PatternFill("solid", fgColor="F4CCCC")
    for row in ws.iter_rows(min_row=2):
        fill = green if row[3].value == "AUTO" else red
        for cell in row:
            cell.fill = fill
    for idx, w in enumerate([15, 65, 65, 22, 26, 12, 26], 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb["Validation Metrics"]
    for idx, w in enumerate([30, 12, 12, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w

    wb.save(output_path)

    # Final Summaryyyy
    total  = len(out_df)
    auto   = counts["AUTO"]
    manual = counts["MANUAL REVIEW NEEDED"]

    print(f"\n{'='*45}")
    print(f"Total processed      : {total}")
    print(f"Auto standardized    : {auto}  ({auto/total*100:.1f}%)")
    print(f"Manual review        : {manual}  ({manual/total*100:.1f}%)")
    print(f"Confidence threshold : {CONFIDENCE_THRESHOLD}")
    print(f"\nModel performance:")
    print(f"  Training accuracy  : {train_acc*100:.2f}%")
    print(f"  Validation accuracy: {val_acc*100:.2f}%")
    print(f"\nOutput: {output_path}")
    print(f"{'='*45}")


if __name__ == "__main__":
    good_path   = sys.argv[1] if len(sys.argv) > 1 else "data/1.Good.xlsx"
    nogood_path = sys.argv[2] if len(sys.argv) > 2 else "data/2.No_Good.xlsx"
    output_path = sys.argv[3] if len(sys.argv) > 3 else "outputs/Classifier_Standardized_Output.xlsx"
    main(good_path, nogood_path, output_path)